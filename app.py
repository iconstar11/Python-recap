import openpyxl as xl

from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(f'{filename}')
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell1 = sheet.cell(1,1)
    # print(cell.value)
    # print(cell1.value)
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        cell_3 = sheet.cell(row, 3)
        print(cell_3.value)
        discount_price = cell_3.value * 0.9
        discount_price_cell = sheet.cell(row, 4)
        discount_price_cell.value = discount_price


    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(f'{filename}2')



'''city_map = {}


cities = ["Nairobi", "Kiambu", "Turkana", "Kisumu"]

city_map["Kenya"] = []
city_map["Tanzania"] = []
city_map["Kenya"] += cities
city_map["Tanzania"] += cities
print(city_map)

print(city_map.values())'''
'''from pathlib import Path
path = Path()


for file in path.glob("*"):
    print(file)
'''

'''import random
class Dice:
    def role_dice(self):
        outcome = random.randint(1,6) , random.randint(1,6)
        return outcome
dice = Dice()
z = dice.role_dice()
print(z)
'''

'''import e_commerce.shipping
from e_commerce.shipping import calc_shipping

calc_shipping()


from utils import findMax

print(findMax([2, 10, 7, 8]))'''


'''import converter
from converter import kg_to_lbs

print(converter.kg_to_lbs(80))
print(kg_to_lbs(80))'''

'''class Mammal:
    def Walk(self):
        print("Walk")

class Cat(Mammal):
    pass

class Dog(Mammal):
    pass

dog1 = Dog()
dog1.Walk()'''


'''class Person:
    def __init__(self, name):
        self.name = name
    def talk(self):
        print("I am speaking and I am")

John = Person("Antony Kinuthia")

John.talk()
print(John.name)'''



'''class Point:
    def __init__(self, w, y):
        self.w = w
        self.y = y
    def move(self):
        print("Move")

    def draw(self):
        print("Draw")


point = Point(10, 20)
print(point.w)'''





'''def square(number):
    return number * number

print(square(13 ))'''



'''try:
    age = int(input("What is your age? "))
    print(age)
except ValueError:
    print("Invalid value")'''


# '''def great_user(name, Id):
#     print(f"Hi {name}, How are You {Id} ?")
#     print("Welcome aboard")
#
#
# print("Start")
# great_user(Id = "John", name = 1)
# print("Finish")
# '''

'''def emoji_converter(sentence):
    words = sentence.split(' ')
    emoji_dict = {
        ":)": "😃",
        "):": "🥲"
    }
    output = ""
    for word in words:
        output += emoji_dict.get(word, word) + " "
    return output


message = input(">")
print(emoji_converter(message))'''





'''phone = input("Phone:")
mapping_dig = {
    "0": "Zero",
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "Four",
    "5": "Five",
    "6": "Six",
    "7": "Seven",
    "8": "Eight",
    "9": "Nine",
}
output = ""
for num in phone:
    output += mapping_dig.get(num, "i") + " "
    if num in mapping_dig:
        print(mapping_dig[num], end=" ")
print()
print(output)'''
'''f_shape = [2,2, 2, 10
    , 7]
i = f_shape[0]
for num in f_shape:
    if num > i:
        i = num

print(i)
'''






'''for x in range(4):
    for y in range(4):
        print(f'x:{x}, y:{y}')'''


'''prices = [10, 20, 30]
total = 0
for price in prices:
    total += price
print(f'Total = {total}')'''


'''for i in range(4):
    print("?", end='')
print()
'''

'''def main():
    height = get_height()
    for i in range(height):
        print("Meow !")


def get_height():
    while True:
        try:
            n = int(input("Height: "))
            if n > 0:
                return n
        except ValueError:
            print("Not a positive integer")


main()'''


'''For Loops'''
'''numbers = [5, 2, 5, 2, 2]
x = 0

for no in numbers:
    output = ''
    for count in range(no):
        output += 'X'

    print(output)
'''
'''prices = [10, 20 , 30]
cost = 0
for item in prices:
    cost += item
print(f'cost is {cost}')
'''

'''for x in range(4):
    for y in range(4):
        print(f'x: {x} and y: {y}')
'''
'''
command = ''
started = False

while True:
    command = input('>_').lower()
    if command == 'start':
        if started == True:
            print('Cannot start a started car')
        else:
            started = True
            print('Car have started....')
    elif command == 'stop':
        if started == False:
            print('Car is not moving')
        else:
            started = False
            print('Car stopped')

    elif command == 'help':
        print(''
        start - To start the car
        stop - To stop the car
        quit - To exit the game       
        '')
    elif command == 'quit':
        break
    else:
        print('Wrong command')


'''

'''
i= 0
correct_number = 9


while i<3:
    your_guess = int(input('Guess a number btwn 1 and 9 '))
    if your_guess == correct_number:
        print(f'You are correct the number is {correct_number}')
        break
    i += 1
else: print('You failed')

'''




'''weight = input('Weight: ')
unit_of_Measure = input('L(bs) or K(g): ')

if unit_of_Measure.upper() == 'K':
    L = int(weight) * 2.2
    print(f'You are {L} pounds')
elif unit_of_Measure.upper() == 'L':
    Kg = int(weight) * 0.45
    print(f'You are {Kg} Kgs')
else: print('Enter a valid unit of measure or weight')'''





'''name = input('Whats your name? ')


if len(name) < 3:
    print('Name must have more than 2 characters')
elif len(name) > 50:
    print('Name cannot have more than 50 characters')
else: print(f'Hi {name}')
'''
'''temp =9

if temp>30:
    print('Its a hot day')
elif temp<10:
    print('Its a cold day')
else: print('Its neither hot or cold')'''




'''good_credit_worth = False

price = 1000000
good_dic = int(price * 90 /100)
bad_cred_dic = int(price * 80 /100)


if good_credit_worth:
    print(f'You need to put down a payment of ${good_dic} as down payment')
else:
    print(f'You need to put down a payment of ${bad_cred_dic} as down payment')

'''
'''is_hot = False
is_cold = False

if is_hot:
    print('Its a hot day')
elif is_cold:
    print('Its a cold day')
else:
    print('Its a lovely day')
print('Enjoy your day')'''


import math

# === Arithmetic Operations ====

'''
x = 10
y = 3
z = 11
z %= 2

print(z)
print(x + y)
print(x - y)
print(x * y)
print(x ** y)
print(x / y)
print(x // y)
print(x % y)
print(x + y)
print(x + y)
'''

#======Math Functions=====

'''z = 2.4

print(abs(z))
print(math.ceil(z))
'''









'''


lang = 'Python'
target = 'beginners '
title = lang + ' course for' + ' [' + target + '] Hi'

heading = f'{lang} course for [{target}]'
print(heading)

print(len(target))
print(target.find('e'))
print(title.title())
'''





# === Hello World ====

'''print('Antony Titan')

print('0----')
print(' IIII')
print('*' * 10)
'''
''' PYTHON VARIABLES '''
'''price = 10
price = 20
rating = 4.5
_isPublished = True
project = 'Python Recap'

print('price', price)
print('rating', rating)
'''

''' Patient Exercise'''
'''name = 'John Smith'
age = 20
_isNewPatient = True'''

'''name = input('What is your name? ')
color = input('What is your favourite colour? ')
print(name + '\'s favourite colour is ' + color)

'''

'''birth_year = input('What is your year of birth? ')

age = 2025 - int(birth_year)

print(age)'''

'''weight_kg = input('What is your weight in kg\'s? ')

weight_pounds = 2.204623 * float(weight_kg)

print('Your weight is', weight_pounds, 'pounds' )'''